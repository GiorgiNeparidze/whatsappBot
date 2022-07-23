import time
import openpyxl
from tkinter import *
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import InvalidArgumentException, TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import ElementNotInteractableException


class account9:
    def Gui9(master):
        account9.costumise9_frame = LabelFrame(master, text='9')

        # Align well
        Grid.columnconfigure(master, 1, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 0, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 1, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 2, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 3, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 4, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 5, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 6, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 7, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 8, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 9, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 10, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 11, weight=1)
        Grid.rowconfigure(account9.costumise9_frame, 12, weight=1)


        def save9(): 
            account9.t1 = textMSG1_for9.get()
            print(account9.t1)
            account9.b1 = banner1_for9.get()
            print(account9.b1)
            account9.c1 = banner1_comment_9.get()
            print(account9.c1)
            account9.b2 = banner2_for9.get()
            print(account9.b2)
            account9.c2 = banner2_comment_9.get()
            print(account9.c2)
            account9.b3 = banner3_for9.get()
            print(account9.b3)
            account9.c3 = banner3_comment_9.get()
            print(account9.c3)
            account9.b4 = banner4_for9.get()
            print(account9.b4)
            account9.c4 = banner4_comment_9.get()
            print(account9.c4)
            account9.b5 = banner5_for9.get()
            print(account9.b5)
            account9.c5 = banner5_comment_9.get()
            print(account9.c5)
            account9.t2 = textMSG2_for9.get()
            print(account9.t2)


        def clear():
            textMSG1_for9.delete(0, END)
            banner1_for9.delete(0, END)
            banner1_comment_9.delete(0, END)
            banner2_for9.delete(0, END)
            banner2_comment_9.delete(0, END)
            banner3_for9.delete(0, END)
            banner3_comment_9.delete(0, END)
            banner4_for9.delete(0, END)
            banner4_comment_9.delete(0, END)
            banner5_for9.delete(0, END)
            banner5_comment_9.delete(0, END)
            textMSG2_for9.delete(0, END)
            del account9.t1
            del account9.b1
            del account9.c1
            del account9.b2
            del account9.c2
            del account9.b3
            del account9.c3
            del account9.b4
            del account9.c4
            del account9.b5
            del account9.c5
            del account9.t2


        def open_path1():
            tv1 = filedialog.askopenfilename()
            banner1_for9.insert(0, tv1)


        def open_path2():
            tv2 = filedialog.askopenfilename()
            banner2_for9.insert(0, tv2)


        def open_path3():
            tv3 = filedialog.askopenfilename()
            banner3_for9.insert(0, tv3)


        def open_path4():
            tv4 = filedialog.askopenfilename()
            banner4_for9.insert(0, tv4)


        def open_path5():
            tv5 = filedialog.askopenfilename()
            banner5_for9.insert(0, tv5)


        # Starting text message
        Label(account9.costumise9_frame, text='Enter Your Starting Text Message: ').grid(row=0, column=0)
        textMSG1_for9 = Entry(account9.costumise9_frame, width=50)
        textMSG1_for9.grid(row=0, column=1)
        # Starting text message
        # Slide 1 > Comment
        Label(account9.costumise9_frame, text="Enter Your First Banner's Name: ").grid(row=1, column=0)
        banner1_for9 = Entry(account9.costumise9_frame, width=50)
        banner1_for9.grid(row=1, column=1)
        Button(account9.costumise9_frame, text='Open', command=open_path1).grid(row=1, column=2)
        Label(account9.costumise9_frame, text='Enter Comment For Your First Banner: ').grid(row=2, column=0)
        banner1_comment_9 = Entry(account9.costumise9_frame, width=50)
        banner1_comment_9.grid(row=2, column=1)
        # Slide 1 > Comment
        # Slide 2 > Comment
        Label(account9.costumise9_frame, text="Enter Your Second Banner's Name: ").grid(row=3, column=0)
        banner2_for9 = Entry(account9.costumise9_frame, width=50)
        banner2_for9.grid(row=3, column=1)
        Button(account9.costumise9_frame, text='Open', command=open_path2).grid(row=3, column=2)
        Label(account9.costumise9_frame, text='Enter Comment For Your Second Banner: ').grid(row=4, column=0)
        banner2_comment_9 = Entry(account9.costumise9_frame, width=50)
        banner2_comment_9.grid(row=4, column=1)
        # Slide 2 > Comment
        # Slide 3 > Comment
        Label(account9.costumise9_frame, text="Enter Your Third Banner's Name: ").grid(row=5, column=0)
        banner3_for9 = Entry(account9.costumise9_frame, width=50)
        banner3_for9.grid(row=5, column=1)
        Button(account9.costumise9_frame, text='Open', command=open_path3).grid(row=5, column=2)
        Label(account9.costumise9_frame, text='Enter Comment For Your Third Banner: ').grid(row=6, column=0)
        banner3_comment_9 = Entry(account9.costumise9_frame, width=50)
        banner3_comment_9.grid(row=6, column=1)
        # Slide 3 > Comment
        # Slide 4 > Comment
        Label(account9.costumise9_frame, text="Enter Your Fourth Banner's Name: ").grid(row=7, column=0)
        banner4_for9 = Entry(account9.costumise9_frame, width=50)
        banner4_for9.grid(row=7, column=1)
        Button(account9.costumise9_frame, text='Open', command=open_path4).grid(row=7, column=2)
        Label(account9.costumise9_frame, text='Enter Comment For Your Fourth Banner: ').grid(row=8, column=0)
        banner4_comment_9 = Entry(account9.costumise9_frame, width=50)
        banner4_comment_9.grid(row=8, column=1)
        # Slide 4 > Comment
        # Slide 5 > Comment
        Label(account9.costumise9_frame, text="Enter Your Fivth Banner's Name: ").grid(row=9, column=0)
        banner5_for9 = Entry(account9.costumise9_frame, width=50)
        banner5_for9.grid(row=9, column=1)
        Button(account9.costumise9_frame, text='Open', command=open_path5).grid(row=9, column=2)
        Label(account9.costumise9_frame, text='Enter Comment For Your Fivth Banner: ').grid(row=10, column=0)
        banner5_comment_9 = Entry(account9.costumise9_frame, width=50)
        banner5_comment_9.grid(row=10, column=1)
        # Slide 5 > Comment
        # Ending text message
        Label(account9.costumise9_frame, text='Enter Your Final Text Massage: ').grid(row=11, column=0)
        textMSG2_for9 = Entry(account9.costumise9_frame, width=50)
        textMSG2_for9.grid(row=11, column=1)
        # Ending text message
            
        # Save & Clear
        sync9_button = Button(account9.costumise9_frame, text='Save', command=save9)
        sync9_button.grid(row=12, column=0, padx=20, pady=20)
        sync9_button = Button(account9.costumise9_frame, text='Clear', command=clear)
        sync9_button.grid(row=12, column=1, padx=20, pady=20)
        


class sync9:
    def __init__(self):
        # PATH to driver
        sync9.PATH9 = r'Engines\9.exe'
        # Driver 10
        sync9.driver9 = webdriver.Chrome(sync9.PATH9)
        # Watsapp web link
        sync9.BASEURL = 'https://web.whatsapp.com/'
        # >>>>> INPUT selector variables <<<<<
        sync9.workbook = openpyxl.load_workbook(r'DATABASE\ac9.xlsx')
        sync9.sheet = sync9.workbook.active
        # ^^^^^ INPUT selector variables ^^^^^
        # >>>>> GET ROW QUANITY <<<<<
        sync9.rows = sync9.sheet.max_row
        sync9.cols = sync9.sheet.max_column
        # ^^^^^ GET COLUMN QUANTITY ^^^^^
        # Starting From
        sync9.rowCount = 2

        # Sync driver usinc BarCode
        def sync_all():
            # Sync
            sync9.driver9.get(sync9.BASEURL)


        sync_all()

        
class send9:
    def __init__(self):
        #Alert Closer driver 9
        def alert_closer(self):
            try:
                sync9.driver9.switch_to.alert.accept()
            except NoAlertPresentException:
                pass
        # Function:  driver9 send msg to needed number
        def watsap_send(self):
            try:
                sync9.driver9.get(send9.URLTOCHAT9)
                continue_t_s  =  WebDriverWait(sync9.driver9, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/a')))
                continue_t_s.click()
                continue_t_ch  =  WebDriverWait(sync9.driver9, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[1]/div[2]/div/div/div[2]/div/div/a')))
                continue_t_ch.click()
                input_div  =  WebDriverWait(sync9.driver9, 30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[2]')))
                # TEXT 1
                try:
                    input_div.send_keys(account9.t1)
                    input_div.send_keys(Keys.RETURN)
                    time.sleep(0.5)
                except InvalidArgumentException:
                    pass
                # Do not cause err
                WebDriverWait(sync9.driver9, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[1]/div[2]/div/div')))
                attachment_box = sync9.driver9.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div/span[2]/div/div[1]/div[2]/div/div')
                attachment_box.click()
                #1>>>
                try:
                    image_box = WebDriverWait(sync9.driver9, 20).until(EC.presence_of_element_located((By.XPATH, '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]')))
                    image_box.send_keys(account9.b1)
                    time.sleep(0.5)
                    img_comment_box = WebDriverWait(sync9.driver9, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/div[1]/div[3]/div/div/div[2]/div[1]/div[2]')))
                    img_comment_box.send_keys(account9.c1)
                    time.sleep(0.5)
                except InvalidArgumentException:
                    pass
                #1^^^
                #2>>>
                try:
                    image_box = WebDriverWait(sync9.driver9, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/input')))
                    image_box.send_keys(account9.b2)
                    time.sleep(0.5)
                    img_comment_box = WebDriverWait(sync9.driver9, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/div[1]/div[3]/div/div/div[2]/div[1]/div[2]')))
                    img_comment_box.send_keys(account9.c2)
                    time.sleep(0.5)
                except InvalidArgumentException:
                    pass
                #2^^^
                #3>>>
                try:
                    image_box = WebDriverWait(sync9.driver9, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/input')))
                    image_box.send_keys(account9.b3)
                    time.sleep(0.5)
                    img_comment_box = WebDriverWait(sync9.driver9, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/div[1]/div[3]/div/div/div[2]/div[1]/div[2]')))
                    img_comment_box.send_keys(account9.c3)
                    time.sleep(0.5)
                except InvalidArgumentException:
                    pass
                #3^^^
                #4>>>
                try:
                    image_box = WebDriverWait(sync9.driver9, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/input')))
                    image_box.send_keys(account9.b4)
                    time.sleep(0.5)
                    img_comment_box = WebDriverWait(sync9.driver9, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/div[1]/div[3]/div/div/div[2]/div[1]/div[2]')))
                    img_comment_box.send_keys(account9.c4)
                    time.sleep(0.5)
                except InvalidArgumentException:
                    pass
                #4^^^
                #5>>>
                try:
                    image_box = WebDriverWait(sync9.driver9, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/input')))
                    image_box.send_keys(account9.b5)
                    time.sleep(0.5)
                    img_comment_box = WebDriverWait(sync9.driver9, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/div[1]/div[3]/div/div/div[2]/div[1]/div[2]')))
                    img_comment_box.send_keys(account9.c5)
                    time.sleep(0.5)
                except InvalidArgumentException:
                    pass
                send_button = sync9.driver9.find_element_by_xpath('//span[@data-icon="send"]')
                send_button.click()
                #5^^^
                try:
                    input_div.send_keys(account9.t2)
                    input_div.send_keys(Keys.RETURN)
                    time.sleep(0.5)
                except InvalidArgumentException:
                    pass
                time.sleep(1)
                # Marksheet
                sync9.sheet.cell(row=sync9.rowCount,column=2).value = 'Inviato'
                sync9.workbook.save(r'DATABASE\ac9.xlsx')
            except TimeoutException:
                sync9.rowCount += 1
                pass

        # To see what a hell is going on
        print('9 Start')

        while sync9.rowCount <= sync9.rows:
            # Selecting needed number for driver 9
            send9.currnetNum = sync9.sheet.cell(row=sync9.rowCount,column=1).value
            send9.URLTOCHAT9 =  'https://wa.me/39'+str(send9.currnetNum)

            # And FINALLY send 
            try:
                watsap_send(self)
            except StaleElementReferenceException:
                pass
            except NoSuchElementException:
                pass
            except ElementClickInterceptedException:
                pass
            except ElementNotInteractableException:
                pass
            except UnexpectedAlertPresentException:
                alert_closer(self)
                continue 
            # Change numberID to next to change to next number automatically
            sync9.rowCount += 1
        
        
        sync9.driver9.close()


class stop9:
    def __init__(self):
        sync9.driver9.close()
